import openpyxl
from openpyxl.utils import get_column_letter
import csv

fields = []
rows = []


def main():
    read_file("lan-devices.csv")
    write_to_excel()

def read_file(file):
    with open(file, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        fields = next(csv_reader)
        for row in csv_reader:
            rows.append(row)
        print(rows[6])

def write_to_excel():
    wb = openpyxl.Workbook()
    sheet = wb.active

    #Add headers to the Excel file
    customer = sheet.cell(row = 1, column = 1)
    customer.value = rows[2][0]
    sheet.column_dimensions['A'].width = 25

    name_header = sheet.cell(row = 1, column = 2)
    name_header.value = rows[2][2]
    sheet.column_dimensions['B'].width = 25

    security_manager_header = sheet.cell(row = 1, column = 3)
    security_manager_header.value = rows[2][9]
    sheet.column_dimensions['C'].width = 25

    security_manager_enabled = sheet.cell(row = 1, column = 4)
    security_manager_enabled.value = "Security Manager Enabled Devices"
    sheet.column_dimensions["D"].width = 30

    security_manager_uninstalled = sheet.cell(row = 1, column = 5)
    security_manager_uninstalled.value = "Security Manager 0.0.0"
    sheet.column_dimensions['E'].width = 25
    
    security_mangager_counter = 0
    security_manager_uninstalled_counter = 0

    for i in range(len(rows)):
        if i >= 3:
            customer = sheet["A%i" % i]
            customer.value = rows[i][0]

            name = sheet["B%i" % i]
            name.value = rows[i][2]

            security_manager = sheet["C%i" % i]
            security_manager.value = rows[i][9]

            if security_manager.value != '--':
                security_mangager_counter += 1
            if security_manager.value == 'Security Manager: 0.0.0':
                security_manager_uninstalled_counter += 1
    
    supposed_to_have = sheet['D3']
    supposed_to_have.value = security_mangager_counter

    uninstalled = sheet['E3']
    uninstalled.value = security_manager_uninstalled_counter


    wb.save(r"C:\Users\Clay Hindman\OneDrive - Invicta Partners LLC\Documents\Code\AV Finder\av_devices.xlsx")


main()